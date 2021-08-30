import os
import pandas as pd
from openpyxl import load_workbook
from .ideb_download import IdebDownload


class ParseIdeb:
    
    def __init__(self, filename, sheet_name, row_ini, row_fim, columns, tipo):
        
        self.download = IdebDownload()
        self.filename = filename
        self.sheet_name = sheet_name
        self.row_ini = row_ini
        self.row_fim= row_fim
        self.columns = columns
        self.tipo = tipo

    def download_data_if_needs(self, filename, tipo):

        if not os.path.exists(filename):
            self.download(tipo=tipo)
        
    def load_wb(self, filename, tipo):
        
        self.download_data_if_needs(filename, tipo)
        wb = load_workbook(filename)
        
        return wb
    
    def get_sheet(self, wb, sheet_name):
        
        return wb[sheet_name]
    
    def parse_data(self, sheet, row_ini, row_fim, columns):
        
        dados = []
        range_dados = range(row_ini, row_fim)
        
        for row in range_dados:    

            line = {col_name : sheet.cell(row, xl_col_id).value for col_name, xl_col_id
                   in columns.items()}
            dados.append(line)
            
        return pd.DataFrame(dados)
    
    def __call__(self):
        
        wb = self.load_wb(self.filename, self.tipo)
        sheet = self.get_sheet(wb, self.sheet_name)
        
        data = self.parse_data(sheet, self.row_ini,
                              self.row_fim, self.columns)
        
        return data

class DataIdebIniciais:
    
    filename = ('raw_data/ideb_raw/divulgacao_anos_iniciais_escolas_2019/'
    'divulgacao_anos_iniciais_escolas_2019.xlsx')
    sheet = 'IDEB_Escolas (Anos_Iniciais)'
    
    columns = dict(
            codigo_municipio = 2,
            codigo_escola = 4,
            nome_escola = 5,
            tipo_rede = 6,
            ideb_2019 = 94
        )

    save_path = 'raw_data/ideb_parsed/'

    row_inicio = 11
    row_fim = 47971
    
    def __init__(self):
        
        self.parser = ParseIdeb(self.filename, 
                                self.sheet,
                                self.row_inicio, 
                                self.row_fim, 
                                self.columns,
                                tipo='iniciais')
        
        self.__data = self.parser()
        self.__data['tipo_anos'] = 'iniciais'

    @property
    def data(self):
        
        return self.__data

    def save_data(self):

        df = self.__data

        if not os.path.exists(self.save_path):
            os.mkdir(self.save_path)

        file = os.path.join(self.save_path, 'dados_ideb_iniciais.csv')

        df.to_csv(file, sep=';', encoding='utf-8')
    
class DataIdebFinais:
    
    filename = ('raw_data/ideb_raw/divulgacao_anos_finais_escolas_2019/'
    'divulgacao_anos_finais_escolas_2019.xlsx')
    sheet = 'IDEB_Escolas (Anos_Finais)'
    
    columns = dict(
            codigo_municipio = 2,
            codigo_escola = 4,
            nome_escola = 5,
            tipo_rede = 6,
            ideb_2019 = 86
        )
    
    save_path = 'raw_data/ideb_parsed/'
    
    row_inicio = 11
    row_fim = 45176
    
    def __init__(self):
        
        self.parser = ParseIdeb(self.filename, 
                                self.sheet,
                                self.row_inicio, 
                                self.row_fim, 
                                self.columns,
                                tipo='finais')
        
        self.__data = self.parser()
        self.__data['tipo_anos'] = 'finais'
    
    @property
    def data(self):
        
        return self.__data

    def save_data(self):

        df = self.__data

        if not os.path.exists(self.save_path):
            os.mkdir(self.save_path)

        file = os.path.join(self.save_path, 'dados_ideb_finais.csv')

        df.to_csv(file, sep=';', encoding='utf-8')